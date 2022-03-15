## XLSX TO CSV
from unittest import case
import openpyxl
import sys
import json
import os.path
import datetime
import string
import random

origem = sys.argv[1]
file_csv_name = sys.argv[2]
type = sys.argv[3]



## opening the xlsx file
xlsx = openpyxl.load_workbook(origem)

## opening the active sheet
sheet = xlsx.active

## getting the data from the sheet
data = sheet.rows

## creating a csv file
if type == '1':
    save_path_csv = 'Files/Type_1/Folder_CSVs'
    save_path_json = 'Files/Type_1/Folder_JSONs'
elif type == '2':
    save_path_csv = 'Files/Type_2/Folder_CSVs'
    save_path_json = 'Files/Type_2/Folder_JSONs'
elif type == '3':
    save_path_csv = 'Files/Type_3/Folder_CSVs'
    save_path_json = 'Files/Type_3/Folder_JSONs'
elif type == '4':
    save_path_csv = 'Files/Type_4/Folder_CSVs'
    save_path_json = 'Files/Type_4/Folder_JSONs'
##completeName = os.path.join(save_path,"data.csv") 
random_name = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(5))
completeName = os.path.join(save_path_csv, random_name + file_csv_name + ".csv")            
  
csv = open(completeName, "w+")

for row in data:
    l = list(row)
    value_to_print=""
    for i in range(len(l)):
        value_to_print += str(l[i].value)+ ','

    csv.write(str(value_to_print) + "\n" )

## close the csv file
csv.close()

# create dict object to store values
json_dic = {}

# create json name fields
first_row = sheet[1]
col_names = []

for col in first_row:
    col_names.append(col.value)

# Iterate the loop to read the cell values
for i in range(1, sheet.max_row):
    json_dic[str(sheet['A'][i].value)] = {}

    counter = 0
    for col in sheet.iter_cols(1, sheet.max_column):
        val = col[i].value

        if type(val) == datetime.datetime:
            val = str(val)

        json_dic[str(sheet['A'][i].value)][col_names[counter]] = val
        counter += 1

json_path = os.path.join(save_path_json, random_name + file_csv_name + ".json")
json_file = open(json_path, 'wb')
print(json_dic)
json_content = json.dumps(json_dic, indent=4).encode('utf-8')
json_file.write(json_content)
json_file.close()

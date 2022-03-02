## XLSX TO CSV
import openpyxl
import sys
import json
import os.path
import datetime

origem = sys.argv[1]
file_csv_name = sys.argv[2]

## opening the xlsx file
xlsx = openpyxl.load_workbook(origem)

## opening the active sheet
sheet = xlsx.active

## getting the data from the sheet
data = sheet.rows

## creating a csv file
save_path_csv = '.Files/Folder_CSVs'
save_path_json = '.Files/Folder_JSONs'

##completeName = os.path.join(save_path,"data.csv") 
completeName = os.path.join(save_path_csv, file_csv_name + ".csv")            
  
csv = open(completeName, "w+")

for row in data:
    l = list(row)
    value_to_print=""
    for i in range(len(l)):
        value_to_print += str(l[i].value)+ ','

    csv.write(str(value_to_print) + "\n" )

## close the csv file
csv.close()

# create dict object to store values
json_dic = {}

# create json name fields
first_row = sheet[1]
col_names = []

for col in first_row:
    col_names.append(col.value)

# Iterate the loop to read the cell values
for i in range(1, sheet.max_row):
    json_dic[str(sheet['A'][i].value)] = {}

    counter = 0
    for col in sheet.iter_cols(1, sheet.max_column):
        val = col[i].value

        if type(val) == datetime.datetime:
            val = str(val)

        json_dic[str(sheet['A'][i].value)][col_names[counter]] = val
        counter += 1

json_file = open(save_path_json+"/"+file_csv_name+".json", 'wb')
json_content = json.dumps(json_dic, indent=4, sort_keys=True).encode('utf-8')
json_file.write(json_content)
json_file.close()